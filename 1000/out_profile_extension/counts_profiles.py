import os
import json
import unicodedata
from collections import Counter
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from io import BytesIO

# paths
INPUT_FOLDER = r"C:/Users/Juan/Desktop/TFM/1000/out_profile_extension/response_clean"
OUTPUT_FOLDER = r"C:/Users/Juan/Desktop/TFM/1000/out_profile_extension/stats"
IMAGES_DIR = r"C:/Users/Juan/Desktop/TFM/1000/out_profile_images_prompts/images_with_ids"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)
OUT_XLSX = os.path.join(OUTPUT_FOLDER, "stats.xlsx")

# canonical lists
ETHNICITY_ALLOWED = [
    "South Asian",
    "East Asian",
    "Southeast Asian",
    "Sub-Saharan African",
    "North African & Middle Eastern (MENA)",
    "European (Northern & Eastern)",
    "European (Southern / Mediterranean)",
    "North American",
    "Central & South American",
    "Oceanian / Pacific Islander",
]

GENDER_ALLOWED = [
    "male",
    "female",
    "non-binary",
    "transgender male",
    "transgender female",
]

ORIENTATION_ALLOWED = [
    "heterosexual",
    "gay/lesbian",
    "bisexual",
    "asexual",
    "pansexual",
    "queer/other",
    "unspecified",  # allowed for minors
]

AGE_BUCKET_LABELS = ["<18", "18-29", "30-39", "40-49", "50-59", "60+"]

# utilities
def strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))


def norm_text(s):
    # Returns (original, normalized_ascii_lowercase).
    if not isinstance(s, str):
        return "", ""
    s = s.strip()
    if not s:
        return "", ""
    s_ascii = strip_accents(s).lower()
    return s, s_ascii


def age_to_bucket(age):
    """Categorizes age into predefined buckets."""
    if not isinstance(age, int):
        return "Unknown"
    if age < 18: return "<18"
    if 18 <= age <= 29: return "18-29"
    if 30 <= age <= 39: return "30-39"
    if 40 <= age <= 49: return "40-49"
    if 50 <= age <= 59: return "50-59"
    return "60+"


def normalize_gender_identity(val):
    """Normalizes gender identity strings to canonical categories."""
    s, a = norm_text(val)
    if not a: return "Unknown"

    # direct matches
    if a in ("male", "m", "man", "masculine"):      return "male"
    if a in ("female", "f", "woman", "feminine"):   return "female"

    # transgender
    import re
    if "trans" in a:
        if re.search(r"\b(man|male|ftm|f2m)\b", a): 
            return "transgender male"
        if re.search(r"\b(woman|female|mtf|m2f)\b", a): 
            return "transgender female"

    # non-binary
    if a in ("non-binary", "nonbinary", "nb", "enby", "genderqueer", "no binario", "nobinario"):
        return "non-binary"
    if "non" in a and "binary" in a:
        return "non-binary"

    if s in GENDER_ALLOWED:
        return s
    return "Unknown"


def normalize_orientation(val):
    """Normalizes sexual orientation strings to canonical categories."""
    s, a = norm_text(val)
    if not a: return "Unknown"

    if a in ("heterosexual", "straight", "hetero"):
        return "heterosexual"
    if a in ("gay", "lesbian", "homosexual", "homo"):
        return "gay/lesbian"
    if a in ("bi", "bisexual"):
        return "bisexual"
    if a in ("asexual", "ace"):
        return "asexual"
    if a in ("pan", "pansexual"):
        return "pansexual"
    if a == "unspecified":
        return "unspecified"
    if "queer" in a or "demi" in a or "fluid" in a or "question" in a or "omni" in a:
        return "queer/other"

    if s in ORIENTATION_ALLOWED:
        return s
    return "Unknown"

# ethnicity normalization (long -> canonical)
LONG_TO_CANON = {
    "south asian (india, pakistan, sri lanka, nepal, etc.)": "South Asian",
    "east asian (china, korea, japan, mongolia, etc.)": "East Asian",
    "southeast asian (vietnam, thailand, philippines, indonesia, etc.)": "Southeast Asian",
    "sub-saharan african (nigeria, kenya, ethiopia, south africa, etc.)": "Sub-Saharan African",
    "north african & middle eastern (mena) (egypt, morocco, saudi arabia, iran, etc.)": "North African & Middle Eastern (MENA)",
    "european (northern & eastern) (germany, poland, sweden, russia, etc.)": "European (Northern & Eastern)",
    "european (southern / mediterranean) (italy, spain, greece, portugal, etc.)": "European (Southern / Mediterranean)",
    "north american (usa, canada, greenland, etc.)": "North American",
    "central & south american (mexico, peru, bolivia, chile, etc.)": "Central & South American",
    "oceanian / pacific islander (fiji, samoa, papua new guinea, hawaii, etc.)": "Oceanian / Pacific Islander",
}
SHORT_LC_TO_CANON = {c.lower(): c for c in ETHNICITY_ALLOWED}


def normalize_ethnicity(val):
    """Normalizes ethnicity strings to canonical categories."""
    s, a = norm_text(val)
    if not a: return "Unknown"

    # exact short match
    if s in ETHNICITY_ALLOWED:
        return s

    # normalize hyphens/spaces
    a = a.replace("–", "-").replace("—", "-")
    a = " ".join(a.split())

    # long → canonical
    if a in LONG_TO_CANON:
        return LONG_TO_CANON[a]

    # short (lowercase) → canonical
    if a in SHORT_LC_TO_CANON:
        return SHORT_LC_TO_CANON[a]

    # heuristic by substring
    if "southeast asian" in a:
        return "Southeast Asian"
    if "east asian" in a and "southeast" not in a:
        return "East Asian"
    if "south asian" in a:
        return "South Asian"
    if "sub-saharan african" in a or "sub saharan african" in a:
        return "Sub-Saharan African"
    if "north african" in a or "middle eastern" in a or "mena" in a:
        return "North African & Middle Eastern (MENA)"
    if "european (northern & eastern)" in a or "northern european" in a or "eastern european" in a:
        return "European (Northern & Eastern)"
    if "european (southern / mediterranean)" in a or "southern european" in a or "mediterranean" in a:
        return "European (Southern / Mediterranean)"
    if "north american" in a:
        return "North American"
    if "central & south american" in a or "central and south american" in a:
        return "Central & South American"
    if "oceanian / pacific islander" in a or "oceanian" in a or "pacific islander" in a:
        return "Oceanian / Pacific Islander"

    return "Unknown"

# country normalization
COUNTRY_ALIASES = {
    # USA
    "us": "United States of America",
    "u.s.": "United States of America",
    "u.s": "United States of America",
    "usa": "United States of America",
    "united states": "United States of America",
    "united states of america": "United States of America",
    "america": "United States of America",
    "ee.uu.": "United States of America",
    # UK
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "united kingdom": "United Kingdom",
    "great britain": "United Kingdom",
    "britain": "United Kingdom",
    "england": "United Kingdom",
    "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "northern ireland": "United Kingdom",
    # Korea
    "korea, republic of": "South Korea",
    "south korea": "South Korea",
    "republic of korea": "South Korea",
    "korea": "South Korea",
    "korea, democratic people's republic of": "North Korea",
    "north korea": "North Korea",
    # common variants
    "viet nam": "Vietnam",
    "czech republic": "Czechia",
    "czechia": "Czechia",
    "turkiye": "Turkey",
    "turkey": "Turkey",
    "cote d'ivoire": "Côte d'Ivoire",
    "cote d’ivoire": "Côte d'Ivoire",
    "ivory coast": "Côte d'Ivoire",
    "russian federation": "Russia",
    "uae": "United Arab Emirates",
    "u.a.e.": "United Arab Emirates",
    "emiratos arabes unidos": "United Arab Emirates",
    "bolivia (plurinational state of)": "Bolivia",
    "brunei darussalam": "Brunei",
    "iran, islamic republic of": "Iran",
    "syrian arab republic": "Syria",
    "moldova, republic of": "Moldova",
    "tanzania, united republic of": "Tanzania",
    "venezuela (bolivarian republic of)": "Venezuela",
    "lao people's democratic republic": "Laos",
    "palestine, state of": "Palestine",
    "macedonia, the former yugoslav republic of": "North Macedonia",
    "north macedonia": "North Macedonia",
    "hong kong": "Hong Kong",
    "taiwan": "Taiwan",
    "myanmar (burma)": "Myanmar",
    "burma": "Myanmar",
    "eswatini": "Eswatini",
    "swaziland": "Eswatini",
}


def normalize_country(val):
    """Normalizes country names to a standard format."""
    s, a = norm_text(val)
    if not a:
        return "Unknown"
    a = a.replace(".", "").strip()
    if a in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[a]
    # leave original capitalization if no alias found
    return s if s else "Unknown"


def counter_to_df(counter: Counter, order: list | None = None) -> pd.DataFrame:
    """Converts a Counter object to a DataFrame with percentages."""
    if order is None:
        items = counter.most_common()
    else: # prioritize ordered keys, then append the rest
        items = [(k, counter.get(k, 0)) for k in order] + [(k, v) for k, v in counter.items() if k not in order]

    total = sum(counter.values()) or 1
    rows = []
    for key, cnt in items:
        pct = cnt / total
        rows.append((key, cnt, pct))

    df = pd.DataFrame(rows, columns=["value", "count", "percent"])
    return df


def autoformat_sheet(ws):
    """Applies auto-width and freezes headers for an Excel sheet."""
    # freeze header
    ws.freeze_panes = "A2"
    # auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = cell.value
            if val is None:
                length = 0
            else:
                s = str(val)
                length = len(s)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def find_image_for(dialogue_id: str, who: str, images_dir: str) -> str | None:
    """
    {dialogue_id}_{who}.png
    {dialogue_id.lower()}_{who}.png
    Accepts extensions: .png .jpg .jpeg .webp
    """
    if not dialogue_id: 
        return None
    bases_to_try = []

    bases_to_try.append(f"{dialogue_id}_{who}")
    bases_to_try.append(f"{dialogue_id.lower()}_{who}")

    exts = [".png",".jpg",".jpeg",".webp"]
    for base in bases_to_try:
        for ext in exts:
            cand = os.path.join(images_dir, base + ext)
            if os.path.isfile(cand):
                return cand
    return None


def make_compressed_image_stream(img_path: str, target_h: int = 160, quality: int = 70) -> BytesIO:
    """
    Opens image, resizes to target_h height (keeping aspect ratio),
    flattens transparency over white if needed, and saves to compressed JPEG
    in a memory buffer. Returns BytesIO ready for openpyxl.Image.
    """
    im = PILImage.open(img_path)

    # Handle transparency (flatten over white)
    if im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info):
        bg = PILImage.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[-1] if im.mode in ("RGBA", "LA") else None)
        im = bg
    else:
        # Ensure RGB for JPEG saving
        if im.mode != "RGB":
            im = im.convert("RGB")

    # Resize to avoid embedding unnecessary megapixels
    if target_h and im.height > target_h:
        ratio = target_h / im.height
        new_size = (max(1, int(im.width * ratio)), target_h)
        im = im.resize(new_size, PILImage.LANCZOS)

    # Save compressed JPEG to memory
    buf = BytesIO()
    im.save(buf, format="JPEG", quality=quality, optimize=True, progressive=True)
    buf.seek(0)
    return buf


def main():
    files = sorted([f for f in os.listdir(INPUT_FOLDER) if f.lower().endswith(".json")])

    # counters
    age_raw = Counter()
    age_bucket = Counter()
    gender_ct = Counter()
    orient_ct = Counter()
    eth_ct = Counter()
    nat_ct = Counter()
    res_ct = Counter()

    # flat list for main sheet
    flat_rows = []
    flat_header = ["dialogue_id", "speaker", "age", "age_bucket",
                   "gender_identity", "sexual_orientation",
                   "ethnicity", "nationality", "residence_country","profile_image"]

    errors = []

    for fname in files:
        fpath = os.path.join(INPUT_FOLDER, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            errors.append((fname, f"load_error: {e}"))
            continue

        did = data.get("dialogue_id", os.path.splitext(fname)[0])
        profiles = data.get("profiles", {})

        for who in ("A", "B"):
            p = profiles.get(who)
            if not isinstance(p, dict):
                continue
            struct = p.get("profile_struct", {})
            if not isinstance(struct, dict):
                continue

            # AGE
            age_val = struct.get("age", None)
            if isinstance(age_val, str):
                try:    age_val = int(age_val.strip())
                except: age_val = None

            bucket = age_to_bucket(age_val)
            age_raw[age_val if isinstance(age_val, int) else "Unknown"] += 1
            age_bucket[bucket] += 1

            # GENDER IDENTITY
            g_val = normalize_gender_identity(struct.get("gender_identity"))
            gender_ct[g_val] += 1

            # ORIENTATION
            so_val = normalize_orientation(struct.get("sexual_orientation"))
            orient_ct[so_val] += 1

            # ETHNICITY
            eth_val = normalize_ethnicity(struct.get("ethnicity"))
            eth_ct[eth_val] += 1

            # NATIONALITY & RESIDENCE
            nat_val = normalize_country(struct.get("nationality"))
            res_val = normalize_country(struct.get("residence_country"))
            nat_ct[nat_val] += 1
            res_ct[res_val] += 1

            # Imagen
            img_path = find_image_for(did, who, IMAGES_DIR)

            # fila plana
            flat_rows.append([
                did,
                who,
                age_val if isinstance(age_val, int) else "",
                bucket,
                g_val,
                so_val,
                eth_val,
                nat_val,
                res_val,
                os.path.basename(img_path) if img_path else ""
            ])

    # create DataFrames
    df_age_raw     = counter_to_df(age_raw)
    df_age_bucket  = counter_to_df(age_bucket, order=AGE_BUCKET_LABELS + ["Unknown"])
    df_gender      = counter_to_df(gender_ct, order=GENDER_ALLOWED + ["Unknown"])
    df_orient      = counter_to_df(orient_ct, order=ORIENTATION_ALLOWED + ["Unknown"])

    # ethnicities: Force the 10 canonical ones to appear, then add others
    eth_order = ETHNICITY_ALLOWED + sorted([k for k in eth_ct.keys() if k not in ETHNICITY_ALLOWED])
    df_eth         = counter_to_df(eth_ct, order=eth_order)
    df_nat         = counter_to_df(nat_ct)   # ordered by frequency
    df_res         = counter_to_df(res_ct)

    df_flat = pd.DataFrame(flat_rows, columns=flat_header)

    # SUMMARY
    total_profiles = len(df_flat)
    summary_rows = [
        ("Total profiles", total_profiles),
        ("Unique nationalities", df_nat.shape[0]),
        ("Unique residence countries", df_res.shape[0]),
        ("Unique ethnicities (observed)", df_eth.shape[0]),
    ]
    df_summary = pd.DataFrame(summary_rows, columns=["metric", "value"])

    # write to Excel
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_age_raw.to_excel(writer, sheet_name="Age_Raw", index=False)
        df_age_bucket.to_excel(writer, sheet_name="Age_Bucket", index=False)
        df_gender.to_excel(writer, sheet_name="Gender_Identity", index=False)
        df_orient.to_excel(writer, sheet_name="Sexual_Orientation", index=False)
        df_eth.to_excel(writer, sheet_name="Ethnicity", index=False)
        df_nat.to_excel(writer, sheet_name="Nationality", index=False)
        df_res.to_excel(writer, sheet_name="Residence_Country", index=False)
        df_flat.to_excel(writer, sheet_name="Profiles_Flat", index=False)

        # formatting: percentages and auto-width
        wb = writer.book

        # sheets with "percent" column (column C)
        for sh in ["Age_Raw", "Age_Bucket", "Gender_Identity", "Sexual_Orientation", "Ethnicity", "Nationality", "Residence_Country"]:
            ws = wb[sh]
            # percent está en la 3ª columna
            for cell in ws["C"][1:]:
                cell.number_format = "0.00%"
            autoformat_sheet(ws)

        # summary and flat sheets
        autoformat_sheet(wb["Summary"])
        ws_flat = wb["Profiles_Flat"]
        autoformat_sheet(ws_flat)

        headers = [c.value for c in ws_flat[1]]
        try:
            img_col_idx = headers.index("profile_image") + 1  # 1-based
        except ValueError:
            img_col_idx = len(headers) + 1

        # increase column width and row height to make thumbnails visible
        img_col_letter = get_column_letter(img_col_idx)
        ws_flat.column_dimensions[img_col_letter].width = 20

        # insert images
        for row_idx in range(2, ws_flat.max_row + 1):
            img_name = ws_flat.cell(row=row_idx, column=img_col_idx).value
            if not img_name:
                continue

            img_path = os.path.join(IMAGES_DIR, img_name)
            if not os.path.isfile(img_path):
                continue

            try:
                # 1) create compressed stream (resize and compress to JPEG)
                jpeg_stream = make_compressed_image_stream(
                    img_path,
                    target_h=160,   # internal height of embedded resource
                    quality=70      # 60–75 good balance
                )

                # 2) create OpenPyXL image from compressed stream
                img = XLImage(jpeg_stream)

                # 3) visual scale in sheet. This only changes display size, not file size.
                display_h = 80
                ratio = display_h / img.height if img.height else 1
                img.height = int(img.height * ratio)
                img.width  = int(img.width  * ratio)

                # 4) adjust row height and anchor
                ws_flat.row_dimensions[row_idx].height = 60
                anchor = f"{img_col_letter}{row_idx}"
                ws_flat.add_image(img, anchor)

            except Exception:
                # if an image fails, skip and continue
                continue

    # console summary
    print(f"Processed files: {len(files)}")
    print(f"Excel created: {OUT_XLSX}")
    if df_flat.empty:
        print("WARNING: No content generated in Profiles_Flat.")

if __name__ == "__main__":
    main()
