import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# excel file
styled_path = "C:/Users/Juan/Desktop/TFM/1000/out_profile_extension/stats/stats.xlsx"
wb = load_workbook(styled_path)

# access sheets
ws_summary = wb["Summary"]
ws_gender = wb["Gender_Identity"]
ws_agebucket = wb["Age_Bucket"]
ws_ethnicity = wb["Ethnicity"]
ws_nationality = wb["Nationality"]
ws_residence = wb["Residence_Country"]
ws_orientation = wb["Sexual_Orientation"]

# clear summary
for row in ws_summary["A2":"Z100"]:
    for cell in row:
        cell.value = None

# add title
ws_summary["A2"] = "Demographic Dashboard"

# Gender Pie Chart
gender_chart = PieChart()
data = Reference(ws_gender, min_col=2, min_row=1, max_row=ws_gender.max_row)
labels = Reference(ws_gender, min_col=1, min_row=2, max_row=ws_gender.max_row)
gender_chart.add_data(data, titles_from_data=True)
gender_chart.set_categories(labels)
gender_chart.title = "Gender Distribution"
ws_summary.add_chart(gender_chart, "B4")

# Age Bucket Bar Chart
age_chart = BarChart()
data = Reference(ws_agebucket, min_col=2, min_row=1, max_row=ws_agebucket.max_row)
cats = Reference(ws_agebucket, min_col=1, min_row=2, max_row=ws_agebucket.max_row)
age_chart.add_data(data, titles_from_data=True)
age_chart.set_categories(cats)
age_chart.title = "Age Buckets"
age_chart.y_axis.title = "Count"
ws_summary.add_chart(age_chart, "L4")

# Ethnicity Bar Chart
eth_chart = BarChart()
data = Reference(ws_ethnicity, min_col=2, min_row=1, max_row=ws_ethnicity.max_row)
cats = Reference(ws_ethnicity, min_col=1, min_row=2, max_row=ws_ethnicity.max_row)
eth_chart.add_data(data, titles_from_data=True)
eth_chart.set_categories(cats)
eth_chart.title = "Ethnicity Distribution"
eth_chart.y_axis.title = "Count"
eth_chart.type = "col"
ws_summary.add_chart(eth_chart, "B20")

# Sexual Orientation Pie Chart
so_chart = PieChart()
data = Reference(ws_orientation, min_col=2, min_row=1, max_row=ws_orientation.max_row)
labels = Reference(ws_orientation, min_col=1, min_row=2, max_row=ws_orientation.max_row)
so_chart.add_data(data, titles_from_data=True)
so_chart.set_categories(labels)
so_chart.title = "Sexual Orientation"
ws_summary.add_chart(so_chart, "L20")

# Top Nationalities Bar Chart
nat_chart = BarChart()
data = Reference(ws_nationality, min_col=2, min_row=1, max_row=10)  # top 10
cats = Reference(ws_nationality, min_col=1, min_row=2, max_row=10)
nat_chart.add_data(data, titles_from_data=True)
nat_chart.set_categories(cats)
nat_chart.title = "Top 10 Nationalities"
nat_chart.y_axis.title = "Count"
nat_chart.type = "bar"
ws_summary.add_chart(nat_chart, "B36")

# Top Residence Countries Bar Chart
res_chart = BarChart()
data = Reference(ws_residence, min_col=2, min_row=1, max_row=10)  # top 10
cats = Reference(ws_residence, min_col=1, min_row=2, max_row=10)
res_chart.add_data(data, titles_from_data=True)
res_chart.set_categories(cats)
res_chart.title = "Top 10 Residence Countries"
res_chart.y_axis.title = "Count"
res_chart.type = "bar"
ws_summary.add_chart(res_chart, "L36")

# save dashboard version
dashboard_path = "C:/Users/Juan/Desktop/TFM/1000/out_profile_extension/stats/stats.xlsx"
wb.save(dashboard_path)
